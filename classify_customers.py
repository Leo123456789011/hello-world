#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Classify customers into L1/L2/L3 and export Excel + TSV."""

from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from explicit_data import EXPLICIT

NAMES_FILE = Path(__file__).with_name("customer_names.txt")


def load_names():
    return [line.strip() for line in NAMES_FILE.read_text(encoding="utf-8").splitlines() if line.strip()]


def build_rows():
    names = load_names()
    rows = []
    missing = []
    for name in names:
        if name not in EXPLICIT:
            missing.append(name)
            rows.append((name, "其他", "其他", ""))
        else:
            l1, l2, l3 = EXPLICIT[name]
            rows.append((name, l1, l2, l3))
    return rows, missing


def export(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "客户分类"
    headers = ["公司名字", "一级分类", "二级分类", "三级分类"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    fills = {
        "金融客户": PatternFill("solid", fgColor="D6EAF8"),
        "产业客户": PatternFill("solid", fgColor="D5F5E3"),
        "个人客户": PatternFill("solid", fgColor="FCF3CF"),
        "其他": PatternFill("solid", fgColor="E5E8E8"),
    }
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    for i, (name, l1, l2, l3) in enumerate(rows, 2):
        ws.cell(i, 1, name).border = thin
        for col, val in enumerate([l1, l2, l3], 2):
            c = ws.cell(i, col, val)
            c.border = thin
            c.alignment = Alignment(horizontal="center")
            if fills.get(l1):
                c.fill = fills[l1]
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.auto_filter.ref = f"A1:D{len(rows) + 1}"
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("分类汇总")
    l1c = Counter(r[1] for r in rows)
    ws2["A1"] = "一级分类"
    ws2["B1"] = "数量"
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font
    ws2["A1"].fill = header_fill
    ws2["B1"].fill = header_fill
    row = 2
    for k in ["金融客户", "产业客户", "个人客户", "其他"]:
        ws2.cell(row, 1, k)
        ws2.cell(row, 2, l1c[k])
        row += 1
    ws2.cell(row, 1, "合计")
    ws2.cell(row, 2, len(rows))

    xlsx = Path("/workspace/客户分类结果.xlsx")
    tsv = Path("/workspace/客户分类结果.tsv")
    wb.save(xlsx)
    with tsv.open("w", encoding="utf-8") as f:
        f.write("公司名字\t一级分类\t二级分类\t三级分类\n")
        for r in rows:
            f.write("\t".join(r) + "\n")
    return xlsx, tsv, l1c


def main():
    rows, missing = build_rows()
    if missing:
        raise SystemExit(f"Missing classifications: {missing}")
    xlsx, tsv, l1c = export(rows)
    print(f"Saved {len(rows)} rows -> {xlsx}, {tsv}")
    for k in ["金融客户", "产业客户", "个人客户", "其他"]:
        print(f"  {k}: {l1c[k]}")


if __name__ == "__main__":
    main()
