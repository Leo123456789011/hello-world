#!/usr/bin/env python3
"""Generate Excel: map products to 事业部 / 产业群 / 项目经理."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from collections import OrderedDict, Counter

# 图二对照
ORG = [
    ("橡塑", "聚烯烃及下游", "韩永"),
    ("橡塑", "C2C3及下游", "孙伟卓"),
    ("橡塑", "橡胶及下游", "王媛媛"),
    ("能化", "芳烃及下游", "李春燕"),
    ("能化", "煤化", "李春燕"),
    ("能化", "炼油", "王能"),
    ("能化", "新能源新材料", "王能"),
    ("钢铁", "钢铁", "初晓"),
    ("钢铁", "富宝", "张超"),
    ("农业", "食品原料及造纸", "毛临江"),
    ("农业", "初级农产品", "王舒同"),
    ("农业", "畜牧及饲料", "牛磊"),
    ("能化", "农资-能化", "李春燕"),
]

# 事业部配色
BU_FILL = {
    "橡塑": "D6EAF8",
    "能化": "FDEBD0",
    "钢铁": "D5D8DC",
    "农业": "D5F5E3",
}
GROUP_FILL = {
    "聚烯烃及下游": "AED6F1",
    "C2C3及下游": "85C1E9",
    "橡胶及下游": "5DADE2",
    "芳烃及下游": "F5CBA7",
    "煤化": "EDBB99",
    "炼油": "E59866",
    "新能源新材料": "F0B27A",
    "农资-能化": "F9E79F",
    "钢铁": "BFC9CA",
    "富宝": "ABB2B9",
    "食品原料及造纸": "ABEBC6",
    "初级农产品": "82E0AA",
    "畜牧及饲料": "58D68D",
}

# 图一商品（保持原顺序，含重复出现）
# (商品, 事业部, 产业群, 项目经理, 分类依据, 原清单所属链条)
ROWS = [
    ("原油（布伦特）", "能化", "炼油", "王能", "国际原油，炼油板块基准原料", "能源原料"),
    ("原油 (WTI)", "能化", "炼油", "王能", "国际原油，炼油板块基准原料", "能源原料"),
    ("原料煤", "能化", "煤化", "李春燕", "煤化工原料煤", "能源原料"),
    ("动力煤", "能化", "煤化", "李春燕", "动力煤/燃料煤，煤化及能源原料", "能源原料"),
    ("天然气 (LNG)", "能化", "炼油", "王能", "油气能源，与炼油/天然气同属能化能源线", "能源原料"),
    ("乙烷", "橡塑", "C2C3及下游", "孙伟卓", "乙烯裂解原料，C2烯烃链", "能源原料"),
    ("丙烷", "橡塑", "C2C3及下游", "孙伟卓", "PDH/裂解原料，C3烯烃链", "能源原料"),
    ("双氧水（27.5%）", "能化", "农资-能化", "李春燕", "无机化工/双氧水，农资能化常规监测品种", "己内酰胺-PA6"),
    ("外购环己酮", "能化", "芳烃及下游", "李春燕", "苯-环己烷-环己酮-己内酰胺链", "己内酰胺-PA6"),
    ("DMAC", "能化", "芳烃及下游", "李春燕", "有机溶剂，常配套化纤/酰胺产业链", "己内酰胺-PA6"),
    ("环己酮", "能化", "芳烃及下游", "李春燕", "芳烃下游，己内酰胺核心中间体", "己内酰胺-PA6"),
    ("液氨", "能化", "农资-能化", "李春燕", "合成氨，氮肥/农资核心品种", "己内酰胺-PA6"),
    ("100%标准硫酸", "能化", "农资-能化", "李春燕", "硫酸，农资能化（酸肥链）", "己内酰胺-PA6"),
    ("液碱", "能化", "煤化", "李春燕", "烧碱/氯碱无机品，归能化煤化（氯碱）", "己内酰胺-PA6"),
    ("煤炭", "能化", "煤化", "李春燕", "煤炭原料，煤化工", "己内酰胺-PA6"),
    ("己内酰胺", "能化", "芳烃及下游", "李春燕", "锦纶原料，芳烃下游核心品种", "己内酰胺-PA6"),
    ("硫磺", "能化", "农资-能化", "李春燕", "制酸原料，硫磺-硫酸-化肥链", "己内酰胺-PA6"),
    ("苯", "能化", "芳烃及下游", "李春燕", "芳烃龙头品种", "己内酰胺-PA6"),
    ("外购98%酸", "能化", "农资-能化", "李春燕", "98%硫酸，农资能化", "己内酰胺-PA6"),
    ("环己烷", "能化", "芳烃及下游", "李春燕", "苯加氢产物，己内酰胺中间体", "己内酰胺-PA6"),
    ("硫酸铵", "能化", "农资-能化", "李春燕", "氮肥，己内酰胺副产亦作化肥流通", "己内酰胺-PA6"),
    ("聚己内酰胺 (PA6)", "能化", "芳烃及下游", "李春燕", "锦纶6/PA6，己内酰胺直接下游", "己内酰胺-PA6"),
    ("高纯氢氧化钠 (32%) 优等品", "能化", "煤化", "李春燕", "高纯烧碱，氯碱无机品", "尼龙66"),
    ("煤炭", "能化", "煤化", "李春燕", "煤炭原料，煤化工", "尼龙66"),
    ("丙烯 一等品", "橡塑", "C2C3及下游", "孙伟卓", "C3烯烃，丙烯腈/环氧丙烷等上游", "尼龙66"),
    ("液氨 一等品", "能化", "农资-能化", "李春燕", "合成氨，氮肥/农资核心品种", "尼龙66"),
    ("硫酸 优等品", "能化", "农资-能化", "李春燕", "硫酸，农资能化（酸肥链）", "尼龙66"),
    ("主产氢氟酸", "能化", "新能源新材料", "王能", "氟化工原料（萤石-氢氟酸）", "尼龙66"),
    ("丁二烯 聚合级", "橡塑", "橡胶及下游", "王媛媛", "合成橡胶核心单体", "尼龙66"),
    ("己二腈", "橡塑", "橡胶及下游", "王媛媛", "丁二烯法己二腈，尼龙66单体（C4橡胶链延伸）", "尼龙66"),
    ("甲基戊二腈", "橡塑", "橡胶及下游", "王媛媛", "己二腈装置联产品", "尼龙66"),
    ("己二胺", "橡塑", "橡胶及下游", "王媛媛", "己二腈加氢产物，尼龙66单体", "尼龙66"),
    ("己二酸 优等品", "能化", "芳烃及下游", "李春燕", "苯/KA油氧化，芳烃下游（AA）", "尼龙66"),
    ("丙烯腈", "橡塑", "C2C3及下游", "孙伟卓", "丙烯氨氧化，C3下游（腈纶原料）", "尼龙66"),
    ("乙腈", "橡塑", "C2C3及下游", "孙伟卓", "丙烯腈装置副产", "尼龙66"),
    ("硫酸铵", "能化", "农资-能化", "李春燕", "氮肥，丙烯腈/己内酰胺等副产", "尼龙66"),
    ("甲基戊二胺", "橡塑", "橡胶及下游", "王媛媛", "甲基戊二腈加氢，己二胺装置联产品", "尼龙66"),
    ("尼龙66", "橡塑", "橡胶及下游", "王媛媛", "己二胺+己二酸聚合，与己二腈同链", "尼龙66"),
    ("环氧丙烷", "橡塑", "C2C3及下游", "孙伟卓", "丙烯下游，HPPO/氯醇法", "环氧丙烷-醇醚"),
    ("丙二醇", "橡塑", "C2C3及下游", "孙伟卓", "环氧丙烷下游", "环氧丙烷-醇醚"),
    ("丙二醇单甲醚", "橡塑", "C2C3及下游", "孙伟卓", "醇醚及酯，C3下游", "环氧丙烷-醇醚"),
    ("丙二醇异单甲醚", "橡塑", "C2C3及下游", "孙伟卓", "丙二醇甲醚异构体，醇醚及酯", "环氧丙烷-醇醚"),
    ("双氧水 (27.5%)", "能化", "农资-能化", "李春燕", "无机化工/双氧水，亦为HPPO原料", "环氧丙烷-醇醚"),
    ("双氧水 (50%)", "能化", "农资-能化", "李春燕", "高浓度双氧水，HPPO及氧化原料", "环氧丙烷-醇醚"),
    ("二乙基蒽醌", "能化", "农资-能化", "李春燕", "蒽醌法双氧水工作液组分", "环氧丙烷-醇醚"),
    ("丙烯", "橡塑", "C2C3及下游", "孙伟卓", "C3烯烃", "环氧丙烷-醇醚"),
    ("甲醇", "能化", "煤化", "李春燕", "煤化工龙头品种（甲醇及下游）", "煤化工-DMC"),
    ("燃料煤", "能化", "煤化", "李春燕", "燃料煤，煤化工能源", "煤化工-DMC"),
    ("原料煤", "能化", "煤化", "李春燕", "煤化工原料煤", "煤化工-DMC"),
    ("硝酸", "能化", "农资-能化", "李春燕", "氮肥/硝酸，农资能化", "煤化工-DMC"),
    ("氢氧化钠", "能化", "煤化", "李春燕", "烧碱，氯碱无机品", "煤化工-DMC"),
    ("液氨", "能化", "农资-能化", "李春燕", "合成氨，氮肥/农资核心品种", "煤化工-DMC"),
    ("硫磺", "能化", "农资-能化", "李春燕", "制酸原料，硫磺-硫酸-化肥链", "煤化工-DMC"),
    ("碳酸二甲酯 (精DMC)", "能化", "新能源新材料", "王能", "精制DMC，电解液溶剂（新能源）", "煤化工-DMC"),
    ("聚酯级乙二醇", "橡塑", "C2C3及下游", "孙伟卓", "乙烯/煤制乙二醇，聚酯级MEG", "BDO-聚酯-可降解"),
    ("工业级乙二醇", "橡塑", "C2C3及下游", "孙伟卓", "工业级乙二醇，C2下游", "BDO-聚酯-可降解"),
    ("硫酸铵 (一型)", "能化", "农资-能化", "李春燕", "氮肥一型硫酸铵", "BDO-聚酯-可降解"),
    ("1,4-丁二醇 (BDO)", "能化", "煤化", "李春燕", "以电石/炔化路线为主的煤化工品种", "BDO-聚酯-可降解"),
    ("对苯二甲酸 (PTA)", "能化", "芳烃及下游", "李春燕", "PX-PTA-聚酯，芳烃下游", "BDO-聚酯-可降解"),
    ("己二酸 (AA)", "能化", "芳烃及下游", "李春燕", "芳烃下游，PBAT/尼龙66单体", "BDO-聚酯-可降解"),
    ("聚对苯二甲酸-己二酸丁二酯 (PBAT)", "能化", "新能源新材料", "王能", "可降解塑料，新材料", "BDO-聚酯-可降解"),
    ("聚对苯二甲酸丁二醇酯 (PBT)", "橡塑", "聚烯烃及下游", "韩永", "工程塑料（PBT）", "BDO-聚酯-可降解"),
    ("四氢呋喃 (THF)", "能化", "煤化", "李春燕", "BDO下游溶剂/氨纶原料PTMEG中间体", "BDO-聚酯-可降解"),
    ("18%氟硅酸", "能化", "新能源新材料", "王能", "氟硅化工中间体", "气凝胶-有机硅"),
    ("38%氟硅酸", "能化", "新能源新材料", "王能", "氟硅化工中间体", "气凝胶-有机硅"),
    ("98%硫酸", "能化", "农资-能化", "李春燕", "硫酸，农资能化（酸肥链）", "气凝胶-有机硅"),
    ("无水氯化氢", "能化", "新能源新材料", "王能", "有机硅/氯硅烷配套原料", "气凝胶-有机硅"),
    ("甲醇", "能化", "煤化", "李春燕", "煤化工龙头品种（甲醇及下游）", "气凝胶-有机硅"),
    ("二氧化碳", "能化", "新能源新材料", "王能", "工业气体；气凝胶超临界干燥介质", "气凝胶-有机硅"),
    ("一甲基三甲氧基硅烷", "能化", "新能源新材料", "王能", "有机硅单体，气凝胶前驱体相关", "气凝胶-有机硅"),
    ("硝酸", "能化", "农资-能化", "李春燕", "氮肥/硝酸，农资能化", "气凝胶-有机硅"),
    ("氨水", "能化", "农资-能化", "李春燕", "合成氨下游，农资能化", "气凝胶-有机硅"),
    ("硫酸", "能化", "农资-能化", "李春燕", "硫酸，农资能化（酸肥链）", "气凝胶-有机硅"),
    ("六甲基二硅氮烷", "能化", "新能源新材料", "王能", "有机硅（HMDS），电子/气凝胶相关", "气凝胶-有机硅"),
    ("硅粉", "能化", "新能源新材料", "王能", "有机硅/气凝胶硅源", "气凝胶-有机硅"),
    ("中高温气凝胶毡", "能化", "新能源新材料", "王能", "气凝胶制品，新能源新材料", "气凝胶-有机硅"),
    ("深冷气凝胶毡", "能化", "新能源新材料", "王能", "气凝胶制品，新能源新材料", "气凝胶-有机硅"),
    ("工业气凝胶毡", "能化", "新能源新材料", "王能", "气凝胶制品，新能源新材料", "气凝胶-有机硅"),
    ("新能源气凝胶毡", "能化", "新能源新材料", "王能", "气凝胶制品，新能源新材料", "气凝胶-有机硅"),
    ("气凝胶涂料", "能化", "新能源新材料", "王能", "气凝胶制品，新能源新材料", "气凝胶-有机硅"),
    ("气凝胶粉", "能化", "新能源新材料", "王能", "气凝胶制品，新能源新材料", "气凝胶-有机硅"),
    ("硅酸酯", "能化", "新能源新材料", "王能", "气凝胶有机硅源（正硅酸酯类）", "气凝胶-有机硅"),
    ("超高分子量聚乙烯", "橡塑", "聚烯烃及下游", "韩永", "聚乙烯高端牌号，聚烯烃", "UHMWPE"),
    ("超高分子量聚乙烯 (C200)", "橡塑", "聚烯烃及下游", "韩永", "UHMWPE牌号，聚烯烃", "UHMWPE"),
    ("超高分子量聚乙烯 (C300)", "橡塑", "聚烯烃及下游", "韩永", "UHMWPE牌号，聚烯烃", "UHMWPE"),
    ("超高分子量聚乙烯 (C400)", "橡塑", "聚烯烃及下游", "韩永", "UHMWPE牌号，聚烯烃", "UHMWPE"),
    ("超高分子量聚乙烯 (C600)", "橡塑", "聚烯烃及下游", "韩永", "UHMWPE牌号，聚烯烃", "UHMWPE"),
    ("超高分子量聚乙烯 (CX300)", "橡塑", "聚烯烃及下游", "韩永", "UHMWPE牌号，聚烯烃", "UHMWPE"),
    ("超高分子量聚乙烯 (CX400)", "橡塑", "聚烯烃及下游", "韩永", "UHMWPE牌号，聚烯烃", "UHMWPE"),
    ("超高分子量聚乙烯 (CF50)", "橡塑", "聚烯烃及下游", "韩永", "UHMWPE牌号，聚烯烃", "UHMWPE"),
    ("超高分子量聚乙烯 (CF100)", "橡塑", "聚烯烃及下游", "韩永", "UHMWPE牌号，聚烯烃", "UHMWPE"),
]


def font(name="微软雅黑", size=11, bold=False, color="000000"):
    return Font(name=name, size=size, bold=bold, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def border():
    s = Side(style="thin", color="7F8C8D")
    return Border(left=s, right=s, top=s, bottom=s)


def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_header(ws, row, cols, hex_color="1F4E79"):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = font(size=11, bold=True, color="FFFFFF")
        cell.fill = fill(hex_color)
        cell.alignment = align()
        cell.border = border()


def style_cell(cell, hex_color=None, h="center", bold=False):
    cell.font = font(bold=bold)
    cell.alignment = align(h=h)
    cell.border = border()
    if hex_color:
        cell.fill = fill(hex_color)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_sheet_classification(wb):
    ws = wb.active
    ws.title = "商品分类明细"
    headers = ["序号", "商品名称", "事业部", "产业群", "项目经理", "分类依据", "原清单所属链条"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers))

    for i, row in enumerate(ROWS, 1):
        name, bu, group, pm, reason, chain = row
        values = [i, name, bu, group, pm, reason, chain]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i + 1, col, val)
            bg = BU_FILL.get(bu, "FFFFFF") if col in (3, 4, 5) else ("F8F9F9" if i % 2 == 0 else "FFFFFF")
            if col == 4:
                bg = GROUP_FILL.get(group, bg)
            h = "left" if col in (2, 6, 7) else "center"
            style_cell(cell, bg, h=h, bold=(col == 5))

    ws.auto_filter.ref = f"A1:G{len(ROWS) + 1}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    for r in range(2, len(ROWS) + 2):
        ws.row_dimensions[r].height = 22
    set_widths(ws, [8, 36, 12, 20, 12, 48, 22])
    ws.auto_filter.ref = f"A1:G{len(ROWS) + 1}"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.oddHeader.left.text = "商品分类明细（按图一顺序）"
    return ws


def add_sheet_unique(wb):
    """Deduplicate by 商品名称, keep first classification."""
    ws = wb.create_sheet("去重商品分类")
    headers = ["序号", "商品名称", "事业部", "产业群", "项目经理", "分类依据", "图一出现次数"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers), "117A65")

    counts = Counter(r[0] for r in ROWS)
    seen = OrderedDict()
    for row in ROWS:
        if row[0] not in seen:
            seen[row[0]] = row

    for i, (name, data) in enumerate(seen.items(), 1):
        _, bu, group, pm, reason, _ = data
        values = [i, name, bu, group, pm, reason, counts[name]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i + 1, col, val)
            bg = BU_FILL.get(bu, "FFFFFF") if col in (3, 4, 5) else ("E8F8F5" if i % 2 == 0 else "FFFFFF")
            if col == 4:
                bg = GROUP_FILL.get(group, bg)
            h = "left" if col in (2, 6) else "center"
            style_cell(cell, bg, h=h, bold=(col == 5))
        ws.row_dimensions[i + 1].height = 22

    ws.auto_filter.ref = f"A1:G{len(seen) + 1}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    set_widths(ws, [8, 36, 12, 20, 12, 48, 16])
    ws.sheet_properties.tabColor = "117A65"
    return ws


def add_sheet_by_manager(wb):
    ws = wb.create_sheet("按项目经理汇总")
    headers = ["项目经理", "事业部", "产业群", "商品名称", "序号（图一）"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers), "6C3483")

    # sort by manager, then group, then original order
    indexed = [(i + 1, r) for i, r in enumerate(ROWS)]
    manager_order = ["韩永", "孙伟卓", "王媛媛", "李春燕", "王能", "初晓", "张超", "毛临江", "王舒同", "牛磊"]

    def key(item):
        idx, r = item
        pm = r[3]
        return (manager_order.index(pm) if pm in manager_order else 99, r[2], idx)

    indexed.sort(key=key)
    for i, (orig_idx, r) in enumerate(indexed, 1):
        name, bu, group, pm, _, _ = r
        values = [pm, bu, group, name, orig_idx]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i + 1, col, val)
            bg = BU_FILL.get(bu, "FFFFFF")
            if col == 3:
                bg = GROUP_FILL.get(group, bg)
            h = "left" if col == 4 else "center"
            style_cell(cell, bg, h=h, bold=(col == 1))
        ws.row_dimensions[i + 1].height = 22

    ws.auto_filter.ref = f"A1:E{len(ROWS) + 1}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    set_widths(ws, [12, 12, 20, 40, 16])
    ws.sheet_properties.tabColor = "6C3483"
    return ws


def add_sheet_count(wb):
    ws = wb.create_sheet("产业群数量统计")
    headers = ["事业部", "产业群", "项目经理", "图一条目数", "去重商品数"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers), "922B21")

    unique = OrderedDict()
    for r in ROWS:
        unique.setdefault(r[0], r)

    row_counts = Counter((r[1], r[2], r[3]) for r in ROWS)
    uniq_counts = Counter((r[1], r[2], r[3]) for r in unique.values())

    r_i = 2
    for bu, group, pm in ORG:
        key = (bu, group, pm)
        n_all = row_counts.get(key, 0)
        n_uniq = uniq_counts.get(key, 0)
        values = [bu, group, pm, n_all, n_uniq]
        for col, val in enumerate(values, 1):
            cell = ws.cell(r_i, col, val)
            bg = GROUP_FILL.get(group, BU_FILL.get(bu, "FFFFFF"))
            style_cell(cell, bg, bold=(col == 3))
        ws.row_dimensions[r_i].height = 22
        r_i += 1

    # totals
    ws.cell(r_i, 1, "合计")
    ws.cell(r_i, 2, "")
    ws.cell(r_i, 3, "")
    ws.cell(r_i, 4, len(ROWS))
    ws.cell(r_i, 5, len(unique))
    for col in range(1, 6):
        style_cell(ws.cell(r_i, col), "1F4E79", bold=True)
        ws.cell(r_i, col).font = font(bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 28
    set_widths(ws, [12, 20, 12, 16, 16])
    ws.sheet_properties.tabColor = "922B21"

    # bar chart of 图一条目数 for groups with count>0
    chart = BarChart()
    chart.type = "col"
    chart.title = "各产业群图一条目数"
    chart.y_axis.title = "条目数"
    chart.x_axis.title = None
    data = Reference(ws, min_col=4, min_row=1, max_row=14)
    cats = Reference(ws, min_col=2, min_row=2, max_row=14)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.legend = None
    chart.width = 18
    chart.height = 9
    ws.add_chart(chart, "A17")
    return ws


def add_sheet_lookup(wb):
    ws = wb.create_sheet("图二对照表")
    headers = ["事业部", "产业群", "项目经理"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, 3, "1A5276")
    for i, (bu, group, pm) in enumerate(ORG, 1):
        for col, val in enumerate((bu, group, pm), 1):
            cell = ws.cell(i + 1, col, val)
            bg = GROUP_FILL.get(group, BU_FILL.get(bu, "FFFFFF"))
            style_cell(cell, bg, bold=(col == 3))
        ws.row_dimensions[i + 1].height = 22

    note_row = 16
    ws.cell(note_row, 1, "分类说明")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
    style_header(ws, note_row, 3, "1A5276")

    notes = [
        "1. 同一商品按品种本身归属分类，不因出现在不同生产装置而改分到不同产业群。",
        "2. 原油、LNG → 能化/炼油（王能）；煤炭、甲醇、BDO、THF、烧碱 → 能化/煤化（李春燕）。",
        "3. 苯、环己酮、己内酰胺、PA6、PTA、己二酸 → 能化/芳烃及下游（李春燕）。",
        "4. 液氨、硫酸、硫磺、硝酸、氨水、硫酸铵、双氧水 → 能化/农资-能化（李春燕）。",
        "5. 乙烷、丙烷、丙烯、环氧丙烷、丙二醇、醇醚、乙二醇、丙烯腈、乙腈 → 橡塑/C2C3及下游（孙伟卓）。",
        "6. 丁二烯及己二腈—己二胺—尼龙66链 → 橡塑/橡胶及下游（王媛媛）。",
        "7. UHMWPE、PBT → 橡塑/聚烯烃及下游（韩永）。",
        "8. 气凝胶及有机硅、氢氟酸、精DMC、PBAT → 能化/新能源新材料（王能）。",
        "9. 图一商品均落在橡塑/能化；钢铁、农业各产业群本次无对应商品。",
        "10. 图一中煤炭、液氨、甲醇、硫酸铵、双氧水（27.5%）、原料煤等重复出现，明细表按原顺序保留，去重表合并。",
    ]
    for j, text in enumerate(notes):
        r = note_row + 1 + j
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cell = ws.cell(r, 1, text)
        style_cell(cell, "F4F6F7" if j % 2 == 0 else "FFFFFF", h="left")
        ws.row_dimensions[r].height = 22
        ws.cell(r, 2).border = border()
        ws.cell(r, 3).border = border()

    ws.row_dimensions[1].height = 28
    set_widths(ws, [14, 22, 70])
    ws.sheet_properties.tabColor = "1A5276"
    return ws


def main():
    wb = Workbook()
    add_sheet_classification(wb)
    add_sheet_unique(wb)
    add_sheet_by_manager(wb)
    add_sheet_count(wb)
    add_sheet_lookup(wb)
    out = "/workspace/商品分类_事业部产业群项目经理.xlsx"
    wb.save(out)
    print(f"saved {out} rows={len(ROWS)} unique={len({r[0] for r in ROWS})}")


if __name__ == "__main__":
    main()
